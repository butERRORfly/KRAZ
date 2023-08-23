import os
import openpyxl
import csv
import io
import pandas as pd

choose_file = input('Введи имя файла с расширением: ')

# Проверка на расширение файла (.csv, .xlsx, .txt)
if '.csv' in choose_file:
    # Убираем расширение файла
    choose_file = choose_file.replace('.csv', '')
    # Конвертация csv в excel
    wb = openpyxl.Workbook()
    ws = wb.active
    with io.open(f'{choose_file}.csv', encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';')
        for row in reader:
            ws.append(row)
    # Сохраняем вновь созданную книгу
    wb.save(f'{choose_file}.xlsx')
    # Считываем данные из книги
    wb = openpyxl.load_workbook(f'{choose_file}.xlsx')
    # Удаляем книгу
    os.remove(f'{choose_file}.xlsx')

if '.xlsx' in choose_file:
    # Убираем расширение файла
    choose_file = choose_file.replace('.xlsx', '')
    # Считываем данные из книги
    wb = openpyxl.load_workbook(f'{choose_file}.xlsx')

if '.txt' in choose_file:
    # Убираем расширение файла
    choose_file = choose_file.replace('.txt', '')
    ex = pd.read_csv(f'{choose_file}.txt', sep='\t', encoding='windows-1251')
    ex.to_excel(f'{choose_file}.xlsx', index=False)
    # считываем данные из книги
    wb = openpyxl.load_workbook(f'{choose_file}.xlsx')

# Получаем доступ к активному листу
ws = wb.active

# Фильтрация по ключам

# Создаем новые листы
good_sheet = wb.create_sheet('Good')
bad_sheet = wb.create_sheet('Bad')
none_sheet = wb.create_sheet('None')
date_sheet = wb.create_sheet('DateChoose')

# Копируем названия заглавных столбцов на каждый новый лист
main_titles_values_good = [good_sheet.append(row) for row in ws.iter_rows(min_row=1, max_row=1, values_only=True)]
main_titles_values_bad = [bad_sheet.append(row) for row in ws.iter_rows(min_row=1, max_row=1, values_only=True)]
main_titles_values_none = [none_sheet.append(row) for row in ws.iter_rows(min_row=1, max_row=1, values_only=True)]
main_titles_values_date = [date_sheet.append(row) for row in ws.iter_rows(min_row=1, max_row=1, values_only=True)]

# Диапазон ячеек, на котором мы ищем флаги "Good", "Bad" и "None"
search_value_range = ws['D']
for i in search_value_range:

    # Обрабатываем флаг "Good"
    if i.value == 'Good':
        # Добавляем значения ячеек на новый лист "Good"
        good_values = [good_sheet.append(row) for row in ws.iter_rows(min_row=i.row, max_row=i.row, values_only=True)]

    # Обрабатываем флаг "Bad"
    if i.value == 'Bad':
        # Добавляем значения ячеек на новый лист "Bad"
        bad_values = [bad_sheet.append(row) for row in ws.iter_rows(min_row=i.row, max_row=i.row, values_only=True)]

    # Обрабатываем флаг "None"
    if i.value is None:
        # Добавляем значения ячеек на новый лист "None"
        none_values = [none_sheet.append(row) for row in ws.iter_rows(min_row=i.row, max_row=i.row, values_only=True)]

# Ввод интересующей нас даты
question = input('Введите дату: ')

# Диапазон ячеек, на котором мы ищем требуемую дату поиска
search_date_range = ws['B']
for el in search_date_range:

    # Проверяем есть ли значение в ячейке
    if el.value is not None:
        # Задаём фильтр поиска по требуемой дате
        filter_date_search = question
        if filter_date_search in el.value:
            # Добавляем значения ячеек на новый лист "DateChoose"
            date_values = [date_sheet.append(row) for row in ws.iter_rows(min_row=el.row, max_row=el.row, values_only=True)]

# Сохраняем наши изменения
wb.save(f'new_file_{choose_file}.xlsx')